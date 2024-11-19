from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string


workbook = load_workbook(filename="dataset_6_56.xlsx")
print(workbook.sheetnames)

sales_pivot_sheet = workbook["Sales_Pivot"]
aurangabad_ledger_sheet = workbook["Aurangabad Ledger"]

# FILL SALES DATA
# ----------------
# start_small_index_src = column_index_from_string("B")
# start_index_destination = column_index_from_string("D")

# while True:
#     aurangabad_ledger_sheet[get_column_letter(start_index_destination) + "4"] = (
#         "=Sales_Pivot!"
#         + get_column_letter(start_small_index_src)
#         + "6:"
#         + get_column_letter(start_small_index_src)
#         + "55"
#     )
#     start_small_index_src += 1
#     if get_column_letter(start_index_destination) == "DP":
#         break
#     start_index_destination += 4

# FILL Inbound Stock DATA
# ----------------
start_small_index_src = column_index_from_string("B")
start_index_destination = column_index_from_string("E")
while True:
    aurangabad_ledger_sheet[get_column_letter(start_index_destination) + "4"] = (
        "=STOCK_TRANSFER_AURANGABAD!"
        + get_column_letter(start_small_index_src)
        + "3:"
        + get_column_letter(start_small_index_src)
        + "52"
    )
    print(
        "=STOCK_TRANSFER_TABLE!"
        + get_column_letter(start_small_index_src)
        + "3:"
        + get_column_letter(start_small_index_src)
        + "52"
    )
    start_small_index_src += 1
    if get_column_letter(start_index_destination) == "DQ":
        break
    start_index_destination += 4


workbook.save("dataset_6_56.xlsx")
